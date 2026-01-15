# trading_show.py
from __future__ import annotations

import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime
from typing import Any, Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
from supabase import create_client, Client


PLUS_TYPES = {"DEPOSIT", "JUAL"}
MINUS_TYPES = {"WITHDRAW", "BELI", "MATERAI"}
SIMPLE_TYPES = {"DEPOSIT", "WITHDRAW", "MATERAI"}

FMT_DATE = "DD-MMM-YY"
FMT_INT = "#,##0"
FMT_MONEY_2 = "#,##0.00"
FMT_GAIN = "#,##0;(#,##0)"

_PNL_ACTIVE_CACHE: Optional[List[Dict[str, Any]]] = None

load_dotenv()


def create_supabase_client() -> Client:
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE")
    if not url or not key:
        raise RuntimeError("Missing env vars: SUPABASE_URL dan/atau SUPABASE_SERVICE_ROLE")
    return create_client(url, key)

def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    return Decimal(str(v))

def _to_date(value) -> Optional[date]:
    """Parse 'YYYY-MM-DD' atau datetime/date jadi date."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except Exception:
            return None
    return None

def _s(v: Any) -> str:
    return "" if v is None else str(v)

def _upper(v: Any) -> str:
    return _s(v).strip().upper()

def _swap_separators_us_to_id(s: str) -> str:
    # "1,000,000.05" -> "1.000.000,05"
    return s.replace(",", "_").replace(".", ",").replace("_", ".")

def _render_table(headers: List[str], rows: List[List[str]], right_cols: set[str]) -> str:
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))

    def line(ch: str = "-") -> str:
        return "+".join(ch * (w + 2) for w in widths)

    def fmt_row(row: List[str]) -> str:
        parts = []
        for i, cell in enumerate(row):
            if headers[i] in right_cols:
                parts.append(" " + cell.rjust(widths[i]) + " ")
            else:
                parts.append(" " + cell.ljust(widths[i]) + " ")
        return "|".join(parts)

    lines: List[str] = []
    lines.append(line("="))
    lines.append(fmt_row(headers))
    lines.append(line("="))
    
    for r in rows:
        lines.append(fmt_row(r))
    lines.append(line("="))
    
    return "\n".join(lines) + "\n"

def _auto_width(ws, pad: int = 2, max_w: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        mx = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            mx = max(mx, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(mx + pad, max_w)

def _normalize_num_id(value: Any) -> Any:
    """
    Normalisasi angka string format Indonesia:
    - "62.605.874,15" -> 62605874.15
    - "62605874,15" -> 62605874.15
    Jika gagal, kembalikan value apa adanya.
    """
    if not isinstance(value, str):
        return value
    s = value.strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return value

def _safe_float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return default
        # handle Indonesian formats: "62.605.874,15" or "62605874,15"
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return default
    return default

def _safe_int(v: Any, default: int = 0) -> int:
    if v is None:
        return default
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, (float, Decimal)):
        return int(Decimal(str(v)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return default
        s = s.replace(".", "").replace(",", "")
        try:
            return int(s)
        except Exception:
            return default
    return default

def _init_widths(headers: List[str]) -> List[int]:
    return [len(h) for h in headers]

def _update_widths(widths: List[int], row_vals: List[Any]) -> None:
    for i, v in enumerate(row_vals):
        if v is None:
            continue
        s = str(v)
        if len(s) > widths[i]:
            widths[i] = len(s)

def _apply_col_widths(ws, widths: List[int], pad: int = 2, max_w: int = 60) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(w + pad, max_w)

def _style_header_row(ws, row: int, headers: List[str], fill: PatternFill, font: Font) -> None:
    center = Alignment(horizontal="center", vertical="center")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center

def _border_thin() -> Border:
    thin = Side(style="thin", color="808080")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_int_id(v: Any) -> str:
    # 1000000 -> "1.000.000"
    d = _to_decimal(v).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    n = int(d)
    return f"{n:,}".replace(",", ".")

def fmt_money_id_2dp(v: Any) -> str:
    # 1000000.05 -> "1.000.000,05"
    d = _to_decimal(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    s = f"{d:,.2f}"  # US style
    return _swap_separators_us_to_id(s)

def fmt_date_dd_mmm_yy(v: Any) -> str:
    # "YYYY-MM-DD" -> "01-Aug-25"
    if v is None or v == "":
        return ""
    if isinstance(v, date):
        dt = v
    else:
        try:
            dt = date.fromisoformat(str(v))
        except Exception:
            return str(v)
    return dt.strftime("%d-%b-%y")

def month_label(yyyy_mm: str) -> str:
    try:
        y = int(yyyy_mm[:4])
        m = int(yyyy_mm[5:7])
        return date(y, m, 1).strftime("%b-%y")
    except Exception:
        return yyyy_mm
    
def fetch_mutasi_transaksi_for_display(client: Client) -> List[Dict[str, Any]]:
    resp = (
        client.table("trans")
        .select(
            "trx_id,created_at,is_active,"
            "trx_date,trx_type,trx_ticker,trx_price,trx_lot,"
            "trx_broker_fee,trx_vat_broker_fee,"
            "trx_exch_fee,trx_vat_exch_fee,"
            "trx_income_tax,trx_total_fee,trx_total"
        )
        .eq("is_active", True)
        .order("trx_date", desc=False)
        .order("created_at", desc=False)
        .order("trx_id", desc=False)
        .execute()
    )

    rows: List[Dict[str, Any]] = resp.data or []

    def sort_key(r: Dict[str, Any]) -> Tuple[str, int, str, int]:
        trx_date = _s(r.get("trx_date"))
        trx_type = _upper(r.get("trx_type"))
        is_materai = 1 if trx_type == "MATERAI" else 0
        created_at = _s(r.get("created_at"))
        trx_id = int(r.get("trx_id") or 0)
        return (trx_date, is_materai, created_at, trx_id)

    rows.sort(key=sort_key)

    saldo = Decimal("0")
    out: List[Dict[str, Any]] = []

    for r in rows:
        trx_type = _upper(r.get("trx_type"))
        total = _to_decimal(r.get("trx_total"))

        if trx_type in PLUS_TYPES:
            saldo += total
        elif trx_type in MINUS_TYPES:
            saldo -= total

        out.append(
            {
                "Tanggal": r.get("trx_date"),
                "Transaksi": r.get("trx_type"),
                "Ticker": r.get("trx_ticker"),
                "Harga": r.get("trx_price"),
                "Lot": r.get("trx_lot"),
                "Broker Fee": r.get("trx_broker_fee"),
                "VAT Broker Fee": r.get("trx_vat_broker_fee"),
                "Exchange Fee": r.get("trx_exch_fee"),
                "VAT Exchange Fee": r.get("trx_vat_exch_fee"),
                "Income Fee": r.get("trx_income_tax"),
                "Total Fee": r.get("trx_total_fee"),
                "Total": r.get("trx_total"),
                "Saldo": saldo,
            }
        )

    return out

def fetch_income_tax_from_pnl_summary(client: Client) -> Decimal:
    total = Decimal("0")
    pnl_rows = fetch_pnl_summary_active(client)
    
    for r in pnl_rows:
        total += _to_decimal(r.get("pnl_income_tax"))
    
    return total

def fetch_pl_from_pnl_summary(client: Client) -> Decimal:
    total = Decimal("0")
    pnl_rows = fetch_pnl_summary_active(client)
    
    for r in pnl_rows:
        total += _to_decimal(r.get("pnl_amount"))
    
    return total

def fetch_pnl_summaries_for_gain(client: Client) -> List[Dict[str, Any]]:
    pnl_rows = fetch_pnl_summary_active(client)

    rows = list(pnl_rows)
    rows.sort(key=lambda r: r.get("pnl_id") or 0)  # pnl_id asc
    rows.sort(key=lambda r: (r.get("pnl_last_sell_date") is None, r.get("pnl_last_sell_date")), reverse=True)
    return rows or []

def fetch_pnl_summary_active(client: Client) -> List[Dict[str, Any]]:
    global _PNL_ACTIVE_CACHE
    if _PNL_ACTIVE_CACHE is not None:
        return _PNL_ACTIVE_CACHE
    
    resp = (
        client.table("pnl_summary")
        .select(
            "pnl_id,pnl_ticker,pnl_trx_ref,"
            "pnl_first_buy_date,pnl_last_sell_date,"
            "pnl_amount,pnl_total_fee,pnl_income_tax"
        )
        .eq("is_active", True)
        .execute()
    )
    
    _PNL_ACTIVE_CACHE = resp.data or []
    return _PNL_ACTIVE_CACHE

def fetch_trans_for_gain(
        client: Client, ticker: str, 
        pnl_trx_ref: Optional[int], 
        first_buy_date: Optional[Any], 
        last_sell_date: Optional[Any]) -> List[Dict[str, Any]]:
    """
    Ambil transaksi BELI/JUAL untuk membentuk 1 journey.
    """
    q = (
        client.table("trans")
        .select(
            "trx_id,created_at,is_active,"
            "trx_date,trx_type,trx_ticker,trx_price,trx_lot,"
            "trx_broker_fee,trx_vat_broker_fee,"
            "trx_exch_fee,trx_vat_exch_fee,"
            "trx_income_tax,trx_total_fee,trx_total,"
            "trx_ref,trx_is_calc_pnl"
        )
        .eq("is_active", True)
        .in_("trx_type", ["BELI", "JUAL"])
        .eq("trx_is_calc_pnl", True)
    )

    if pnl_trx_ref is not None:
        # (trx_id = root) OR (trx_ref = root)
        # postgrest "or" syntax: "col.eq.value,col.eq.value"
        q = q.or_(f"trx_id.eq.{pnl_trx_ref},trx_ref.eq.{pnl_trx_ref}")
    else:
        q = q.eq("trx_ticker", ticker.strip().upper())
        if first_buy_date:
            q = q.gte("trx_date", str(first_buy_date))
        if last_sell_date:
            q = q.lte("trx_date", str(last_sell_date))

    resp = (
        q.order("trx_date", desc=False)
         .order("created_at", desc=False)
         .order("trx_id", desc=False)
         .execute()
    )
    rows: List[Dict[str, Any]] = resp.data or []

    # Safety filter: pastikan tetap 1 ticker (jaga-jaga kalau query OR mengembalikan data lain)
    t = ticker.strip().upper()
    rows = [r for r in rows if _upper(r.get("trx_ticker")) == t]

    return rows

def fetch_trans_for_gain_batch(client: Client, refs: List[int]) -> List[Dict[str, Any]]:
    if not refs:
        return []

    all_rows: List[Dict[str, Any]] = []
    CHUNK = 200

    for i in range(0, len(refs), CHUNK):
        part = refs[i:i+CHUNK]
        csv = ",".join(str(x) for x in part)

        q = (
            client.table("trans")
            .select(
                "trx_id,created_at,is_active,"
                "trx_date,trx_type,trx_ticker,trx_price,trx_lot,"
                "trx_broker_fee,trx_vat_broker_fee,"
                "trx_exch_fee,trx_vat_exch_fee,"
                "trx_income_tax,trx_total_fee,trx_total,"
                "trx_ref,trx_is_calc_pnl"
            )
            .eq("is_active", True)
            .in_("trx_type", ["BELI", "JUAL"])
            .eq("trx_is_calc_pnl", True)
            .or_(f"trx_id.in.({csv}),trx_ref.in.({csv})")
            .order("trx_date", desc=False)
            .order("created_at", desc=False)
            .order("trx_id", desc=False)
            .execute()
        )
        all_rows.extend(q.data or [])

    return all_rows

def fetch_ticker_pl_from_pnl_summary(client: Client) -> List[Tuple[str, float, int]]:
    pnl_rows = fetch_pnl_summary_active(client)
    agg: Dict[str, Dict[str, float]] = {}

    for r in pnl_rows:
        ticker = (r.get("pnl_ticker") or "").strip().upper()
        if not ticker:
            continue
        pl = _to_decimal(r.get("pnl_amount"))
        qty_inc = 1 if r.get("pnl_trx_ref") is not None else 0

        bucket = agg.get(ticker)
        if bucket is None:
            bucket = {"pl": Decimal("0"), "qty": 0}
            agg[ticker] = bucket
        
        bucket["pl"] += pl
        bucket["qty"] += qty_inc

    out: List[Tuple[str, float, int]] = [
        (t, float(v["pl"]), int(v["qty"])) for t, v in agg.items()
    ]
    out.sort(key=lambda x: x[1], reverse=True)
    return out

def fetch_pnl_summary_for_year(client: Client, year: int) -> List[Dict[str, Any]]:
    start = f"{year:04d}-01-01"
    end = f"{year + 1:04d}-01-01"
    pnl_rows = fetch_pnl_summary_active(client)

    out: List[Dict[str, Any]] = []
    for r in pnl_rows:
        d = r.get("pnl_last_sell_date")
        if not d:
            continue

        ds = d if isinstance(d, str) else str(d)

        if start <= ds < end:
            out.append(r)

    return out

def build_rows_minimal(out: List[Dict[str, Any]]) -> Tuple[List[str], List[List[str]]]:
    headers = ["Tanggal", "Transaksi", "Ticker", "Harga", "Lot", "Total", "Saldo"]
    rows: List[List[str]] = []

    for r in out:
        trx_type = _upper(r["Transaksi"])
        is_simple = trx_type in SIMPLE_TYPES

        rows.append(
            [
                fmt_date_dd_mmm_yy(r["Tanggal"]),
                _s(r["Transaksi"]),
                "" if is_simple else _s(r["Ticker"]),
                "" if is_simple else fmt_int_id(r["Harga"]),
                "" if is_simple else fmt_int_id(r["Lot"]),
                fmt_int_id(r["Total"]),
                fmt_int_id(r["Saldo"]),
            ]
        )

    return headers, rows

def build_rows_full(out: List[Dict[str, Any]]) -> Tuple[List[str], List[List[str]]]:
    headers = [
        "Tanggal", "Transaksi", "Ticker", "Harga", "Lot",
        "Broker Fee", "VAT Broker Fee", "Exchange Fee", "VAT Exchange Fee",
        "Income Fee", "Total Fee", "Total", "Saldo"
    ]
    rows: List[List[str]] = []

    for r in out:
        trx_type = _upper(r["Transaksi"])
        is_simple = trx_type in SIMPLE_TYPES

        rows.append(
            [
                fmt_date_dd_mmm_yy(r["Tanggal"]),
                _s(r["Transaksi"]),
                "" if is_simple else _s(r["Ticker"]),
                "" if is_simple else fmt_int_id(r["Harga"]),
                "" if is_simple else fmt_int_id(r["Lot"]),
                "" if is_simple else fmt_money_id_2dp(r["Broker Fee"]),
                "" if is_simple else fmt_money_id_2dp(r["VAT Broker Fee"]),
                "" if is_simple else fmt_money_id_2dp(r["Exchange Fee"]),
                "" if is_simple else fmt_money_id_2dp(r["VAT Exchange Fee"]),
                "" if is_simple else fmt_money_id_2dp(r["Income Fee"]),
                "" if is_simple else fmt_money_id_2dp(r["Total Fee"]),
                fmt_int_id(r["Total"]),
                fmt_int_id(r["Saldo"]),
            ]
        )

    return headers, rows

def build_trade_performance_metrics(client: Client) -> List[Tuple[str, str]]:
    pnl_rows = fetch_pnl_summary_active(client)

    trade_count = 0
    win_count = 0
    total_pl = Decimal("0")
    total_profit = Decimal("0")
    total_loss_abs = Decimal("0")  # abs(loss)
    holding_days_list: List[int] = []

    for r in pnl_rows:
        trade_count += 1

        pnl = _to_decimal(r.get("pnl_amount"))
        total_pl += pnl
        if pnl > 0:
            win_count += 1
            total_profit += pnl
        elif pnl < 0:
            total_loss_abs += (-pnl)

        d1 = _to_date(r.get("pnl_first_buy_date"))
        d2 = _to_date(r.get("pnl_last_sell_date"))

        if d1 and d2:
            hd = max((d2 - d1).days, 0)
        else:
            hd = 0
        
        holding_days_list.append(hd)

    # Win rate
    if trade_count > 0:
        win_rate = (Decimal(win_count) / Decimal(trade_count)) * Decimal("100")
    else:
        win_rate = Decimal("0")

    # Profit factor
    profit_factor: Optional[Decimal]
    if total_loss_abs == 0:
        profit_factor = None  # "N/A"
    else:
        profit_factor = total_profit / total_loss_abs

    # Avg / Max holding days
    if holding_days_list:
        avg_holding = Decimal(sum(holding_days_list)) / Decimal(len(holding_days_list))
        max_holding = max(holding_days_list)
    else:
        avg_holding = Decimal("0")
        max_holding = 0

    # Format percent: "61,5%"
    win_rate_str = f"{win_rate.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}%".replace(".", ",")

    # Format profit factor: "1,30"
    if profit_factor is None:
        profit_factor_str = "N/A"
    else:
        profit_factor_str = f"{profit_factor.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)}".replace(".", ",")

    # Format avg holding: "14,9"
    avg_holding_str = f"{avg_holding.quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)}".replace(".", ",")

    return [
        ("Jumlah trade (sell events)", fmt_int_id(trade_count)),
        ("Win rate", win_rate_str),
        ("Total realized P/L", float(total_pl.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))),
        ("Total profit", float(total_profit.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))),
        ("Total loss", float(total_loss_abs.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))),
        ("Profit factor", profit_factor_str),
        ("Avg holding days", avg_holding_str),
        ("Max holding days", fmt_int_id(max_holding)),
    ]

def build_open_positions_from_trans(out: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Sheet 3 — Open Positions
    """
    # per ticker: list of chunks FIFO: (lot, cost_total, buy_date)
    # lot = lot (bukan share), cost_total = trx_total (Rp) untuk chunk tersebut
    inv: Dict[str, List[Dict[str, Any]]] = {}

    def parse_date(v: Any) -> Optional[date]:
        if not v:
            return None
        if isinstance(v, date):
            return v
        try:
            return date.fromisoformat(str(v))
        except Exception:
            return None

    for r in out:
        ttype = _upper(r.get("Transaksi"))
        ticker = _s(r.get("Ticker")).strip()
        if not ticker:
            continue

        trx_date = parse_date(r.get("Tanggal"))
        lot = int(_to_decimal(r.get("Lot")))
        if lot <= 0:
            continue

        if ttype == "BELI":
            cost = _to_decimal(r.get("Total"))  # net cash out (incl fee)
            inv.setdefault(ticker, []).append(
                {"lot": lot, "cost": cost, "date": trx_date}
            )

        elif ttype == "JUAL":
            sell_lot = lot
            chunks = inv.get(ticker, [])
            i = 0
            while sell_lot > 0 and i < len(chunks):
                ch = chunks[i]
                ch_lot = int(ch["lot"])
                if ch_lot <= 0:
                    i += 1
                    continue

                if sell_lot >= ch_lot:
                    # consume full chunk
                    sell_lot -= ch_lot
                    i += 1
                else:
                    # partial consume chunk: reduce lot & cost proportionally
                    ratio = Decimal(sell_lot) / Decimal(ch_lot)
                    ch["cost"] = (ch["cost"] * (Decimal("1") - ratio)).quantize(Decimal("0.01"))
                    ch["lot"] = ch_lot - sell_lot
                    sell_lot = 0

            # drop consumed chunks (those before i)
            if i > 0:
                inv[ticker] = chunks[i:]

            # bersihkan chunk yang lot-nya 0
            inv[ticker] = [c for c in inv[ticker] if int(c["lot"]) > 0]

    rows: List[Dict[str, Any]] = []
    for ticker, chunks in inv.items():
        open_lot = sum(int(c["lot"]) for c in chunks)
        if open_lot <= 0:
            continue

        est_cost = sum(_to_decimal(c["cost"]) for c in chunks)  # Rp total
        # Avg cost per share = est_cost / (open_lot * 100)
        denom = Decimal(open_lot) * Decimal("100")
        avg_cost = (est_cost / denom) if denom != 0 else Decimal("0")

        # First/Last buy date dari chunk yang masih tersisa
        dates = [c["date"] for c in chunks if c.get("date")]
        first_buy = min(dates) if dates else None
        last_buy = max(dates) if dates else None

        rows.append(
            {
                "Ticker": ticker,
                "Open Lot": fmt_int_id(open_lot),
                "Avg Cost": float(avg_cost.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Est Cost": float(est_cost.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "First Buy": first_buy,
                "Last Buy": last_buy,
            }
        )

    # Sort: terbesar dulu by Est Cost
    rows.sort(key=lambda x: x["Est Cost"], reverse=True)
    return rows

def build_best_trades_by_ticker(client: Client, top_n: int = 10) -> List[Dict[str, Any]]:
    pnl_rows = fetch_pnl_summary_active(client)

    agg: Dict[str, Dict[str, Any]] = {}
    for r in pnl_rows:
        ticker = _s(r.get("pnl_ticker")).strip().upper()
        if not ticker:
            continue

        pnl = _to_decimal(r.get("pnl_amount"))

        # holding days
        d1 = _to_date(r.get("pnl_first_buy_date"))
        d2 = _to_date(r.get("pnl_last_sell_date"))
        hd = max((d2 - d1).days, 0) if (d1 and d2) else 0

        a = agg.setdefault(
            ticker,
            {"pnl_sum": Decimal("0"), "hold_sum": Decimal("0"), "cnt": 0},
        )
        a["pnl_sum"] += pnl
        a["hold_sum"] += Decimal(hd)
        a["cnt"] += 1

    out: List[Dict[str, Any]] = []
    for ticker, a in agg.items():
        pnl_sum: Decimal = a["pnl_sum"]
        if pnl_sum <= 0:
            continue

        cnt = int(a["cnt"]) if a["cnt"] else 0
        avg_hold = (a["hold_sum"] / Decimal(cnt)) if cnt > 0 else Decimal("0")
        avg_hold = avg_hold.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)

        out.append(
            {
                "Ticker": ticker,
                "Realized P/L": float(pnl_sum.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Avg Holding Period (d)": float(avg_hold),
            }
        )

    out.sort(key=lambda x: x["Realized P/L"], reverse=True)
    return out[:top_n]

def build_worst_trades_by_ticker(client: Client, top_n: int = 10) -> List[Dict[str, Any]]:
    pnl_rows = fetch_pnl_summary_active(client)

    agg: Dict[str, Dict[str, Any]] = {}
    for r in pnl_rows:
        ticker = _s(r.get("pnl_ticker")).strip().upper()
        if not ticker:
            continue

        pnl = _to_decimal(r.get("pnl_amount"))

        # holding days
        d1 = _to_date(r.get("pnl_first_buy_date"))
        d2 = _to_date(r.get("pnl_last_sell_date"))
        hd = max((d2 - d1).days, 0) if (d1 and d2) else 0

        a = agg.setdefault(
            ticker,
            {"pnl_sum": Decimal("0"), "hold_sum": Decimal("0"), "cnt": 0},
        )
        a["pnl_sum"] += pnl
        a["hold_sum"] += Decimal(hd)
        a["cnt"] += 1

    out: List[Dict[str, Any]] = []
    for ticker, a in agg.items():
        pnl_sum: Decimal = a["pnl_sum"]
        cnt = int(a["cnt"]) if a["cnt"] else 0
        avg_hold = (a["hold_sum"] / Decimal(cnt)) if cnt > 0 else Decimal("0")
        avg_hold = avg_hold.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)

        out.append(
            {
                "Ticker": ticker,
                "Realized P/L": float(pnl_sum.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Avg Holding Period (d)": float(avg_hold),
            }
        )

    out.sort(key=lambda x: x["Realized P/L"])
    return out[:top_n]

def build_active_trading_daily_monthly(
        out: List[Dict[str, Any]], 
        client: Client, 
        top_n_days: int = 10) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    pnl_rows = fetch_pnl_summary_active(client)
    realized_by_date: Dict[str, Decimal] = {}
    for r in pnl_rows:
        d = _s(r.get("pnl_last_sell_date"))
        if not d:
            continue
        realized_by_date[d] = realized_by_date.get(d, Decimal("0")) + _to_decimal(r.get("pnl_amount"))

    daily: Dict[str, Dict[str, Any]] = {}

    for r in out:
        d = _s(r.get("Tanggal"))
        if not d:
            continue
        ttype = _upper(r.get("Transaksi"))

        rec = daily.setdefault(
            d,
            {
                "Tanggal": d,
                "#Trx": 0,
                "Turnover": Decimal("0"),
                "Net CF": Decimal("0"),
                "Fee": Decimal("0"),
            },
        )

        rec["#Trx"] += 1

        total = _to_decimal(r.get("Total"))
        if ttype in {"DEPOSIT", "JUAL"}:
            rec["Net CF"] += total
        elif ttype in {"WITHDRAW", "BELI", "MATERAI"}:
            rec["Net CF"] -= total

        if ttype in {"BELI", "JUAL"}:
            price = _to_decimal(r.get("Harga"))
            lot = _to_decimal(r.get("Lot"))
            rec["Turnover"] += price * lot * Decimal("100")
            rec["Fee"] += _to_decimal(r.get("Total Fee"))

    daily_rows: List[Dict[str, Any]] = []
    for d, rec in daily.items():
        rp = realized_by_date.get(d, Decimal("0"))
        daily_rows.append(
            {
                "Tanggal": d,
                "#Trx": rec["#Trx"],
                "Turnover": float(rec["Turnover"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Net CF": float(rec["Net CF"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Realized P/L": float(rp.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Fee": float(rec["Fee"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            }
        )

    daily_rows.sort(key=lambda x: x["Turnover"], reverse=True)
    daily_top = daily_rows[:top_n_days]

    monthly: Dict[str, Dict[str, Any]] = {}
    for row in daily_rows:
        yyyy_mm = row["Tanggal"][:7]  # "YYYY-MM"
        m = monthly.setdefault(
            yyyy_mm,
            {"#Trx": 0, "Turnover": 0.0, "Net CF": 0.0, "Realized P/L": 0.0, "Fee": 0.0},
        )
        m["#Trx"] += row["#Trx"]
        m["Turnover"] += row["Turnover"]
        m["Net CF"] += row["Net CF"]
        m["Realized P/L"] += row["Realized P/L"]
        m["Fee"] += row["Fee"]

    monthly_rows: List[Dict[str, Any]] = []
    for yyyy_mm, m in sorted(monthly.items(), key=lambda kv: kv[0]):
        monthly_rows.append(
            {
                "Bulan": month_label(yyyy_mm),
                "#Trx": m["#Trx"],
                "Turnover": m["Turnover"],
                "Net CF": m["Net CF"],
                "Realized P/L": m["Realized P/L"],
                "Fee": m["Fee"],
            }
        )

    return daily_top, monthly_rows

def build_gain_journey_rows(
        trans_rows: List[Dict[str, Any]], 
        *, 
        pnl_amount: Optional[Any] = None) -> Tuple[List[List[Any]], Dict[str, Decimal]]:
    """
    Convert trans rows -> excel rows + totals.

    Kolom:
    Tanggal, Transaksi, Ticker, Harga, Lot, 
    Broker Fee, VAT Broker Fee, Exchange Fee, VAT Exchange Fee, Income Tax, Total Fee, Jumlah
    """
    rows: List[List[Any]] = []

    broker_sum = Decimal("0")
    vat_broker_sum = Decimal("0")
    exch_sum = Decimal("0")
    vat_exch_sum = Decimal("0")
    tax_sum = Decimal("0")
    fee_sum = Decimal("0")
    buy_sum = Decimal("0")
    sell_sum = Decimal("0")

    for r in trans_rows:
        ttype = _upper(r.get("trx_type"))
        d = date.fromisoformat(str(r.get("trx_date"))) or r.get("trx_date")

        price = _to_decimal(r.get("trx_price"))
        lot = _to_decimal(r.get("trx_lot"))

        broker_fee = _to_decimal(r.get("trx_broker_fee"))
        vat_broker_fee = _to_decimal(r.get("trx_vat_broker_fee"))
        exch_fee = _to_decimal(r.get("trx_exch_fee"))
        vat_exch_fee = _to_decimal(r.get("trx_vat_exch_fee"))
        income_tax = _to_decimal(r.get("trx_income_tax"))

        total_fee = _to_decimal(r.get("trx_total_fee"))
        total = _to_decimal(r.get("trx_total"))

        broker_sum += broker_fee
        vat_broker_sum += vat_broker_fee
        exch_sum += exch_fee
        vat_exch_sum += vat_exch_fee
        tax_sum += income_tax
        fee_sum += total_fee

        if ttype == "BELI":
            buy_sum += total
        elif ttype == "JUAL":
            sell_sum += total

        rows.append([
            d,
            "Beli" if ttype == "BELI" else "Jual",
            _upper(r.get("trx_ticker")),
            int(price.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
            int(lot.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
            float(broker_fee.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(vat_broker_fee.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(exch_fee.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(vat_exch_fee.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(income_tax.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            int(total_fee.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
            int(total.quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
        ])

    gain_calc = sell_sum - buy_sum
    pnl_amount_dec = _to_decimal(pnl_amount) if pnl_amount is not None else None
    gain = pnl_amount_dec if pnl_amount_dec is not None else gain_calc

    totals = {
        "broker_sum": broker_sum,
        "vat_broker_sum": vat_broker_sum,
        "exch_sum": exch_sum,
        "vat_exch_sum": vat_exch_sum,
        "tax_sum": tax_sum,
        "fee_sum": fee_sum,
        "buy_sum": buy_sum,
        "sell_sum": sell_sum,
        "gain": gain,
        "gain_calc": gain_calc,
    }
    return rows, totals

def build_gain_blocks(client: Client) -> List[Dict[str, Any]]:
    pnl_rows = fetch_pnl_summaries_for_gain(client)
    refs: List[int] = []
    pnl_by_ref: Dict[int, Dict[str, Any]] = {}

    for p in pnl_rows:
        ref = p.get("pnl_trx_ref")
        if ref is None:
            continue
        try:
            ref_i = int(ref)
        except Exception:
            continue
        refs.append(ref_i)
        pnl_by_ref[ref_i] = p

    if not refs:
        return []

    trans_rows = fetch_trans_for_gain_batch(client, refs)
    trans_by_root: Dict[int, List[Dict[str, Any]]] = {}
    for t in trans_rows:
        root = t.get("trx_ref") or t.get("trx_id")
        try:
            root_i = int(root)
        except Exception:
            continue
        trans_by_root.setdefault(root_i, []).append(t)

    blocks: List[Dict[str, Any]] = []

    for p in pnl_rows:
        ref = p.get("pnl_trx_ref")
        try:
            ref_i = int(ref) if ref is not None else None
        except Exception:
            ref_i = None
        if ref_i is None:
            continue

        rows = trans_by_root.get(ref_i, [])
        if not rows:
            continue

        pnl_ticker = _upper(p.get("pnl_ticker"))
        if pnl_ticker:
            rows = [r for r in rows if _upper(r.get("trx_ticker")) == pnl_ticker]
            if not rows:
                continue

        rows.sort(
            key=lambda r: (
                _s(r.get("trx_date")),
                _s(r.get("created_at")),
                int(r.get("trx_id") or 0),
            )
        )

        excel_rows, totals = build_gain_journey_rows(rows, pnl_amount=p.get("pnl_amount"))

        blocks.append(
            {
                "rows": excel_rows,
                "totals": totals,
                "pnl_trx_ref": ref_i,
                "pnl_ticker": pnl_ticker,
                "pnl_id": p.get("pnl_id"),
                "pnl_first_buy_date": p.get("pnl_first_buy_date"),
                "pnl_last_sell_date": p.get("pnl_last_sell_date"),
            }
        )

    return blocks

def build_yearly_performance_summary(client: Client, out_trans: List[Dict[str, Any]], year: int) -> Dict[str, Any]:
    pnl_rows = fetch_pnl_summary_for_year(client, year)

    daily_pl: Dict[date, float] = {}
    daily_fee: Dict[date, float] = {}
    wins = 0
    total_trades = 0
    profit = 0.0
    loss = 0.0

    for r in pnl_rows:
        d =  date.fromisoformat(r.get("pnl_last_sell_date"))
        if not d:
            continue
        pl = float(r.get("pnl_amount") or 0)
        fee = float(r.get("pnl_total_fee") or 0)
        daily_pl[d] = daily_pl.get(d, 0.0) + pl
        daily_fee[d] = daily_fee.get(d, 0.0) + fee

        total_trades += 1
        if pl > 0:
            wins += 1
            profit += pl
        elif pl < 0:
            loss += pl  # tetap negatif

    total_net_pl = sum(daily_pl.values())
    total_fee = sum(daily_fee.values())
    best_day = max(daily_pl.values()) if daily_pl else 0.0
    worst_day = min(daily_pl.values()) if daily_pl else 0.0
    win_rate = (wins / total_trades) if total_trades else 0.0

    def _get_trans_date(t: Dict[str, Any]) -> Optional[date]:
        # support raw trans + display trans
        return _to_date(t.get("trx_date") or t.get("Tanggal"))

    def _get_trans_type(t: Dict[str, Any]) -> str:
        return (t.get("trx_type") or t.get("Transaksi") or "").upper()

    def _get_trans_total(t: Dict[str, Any]) -> float:
        v = t.get("trx_total")
        if v is None:
            v = t.get("Total")
        return float(v or 0)
    
    # --- Modal/return: DEPOSIT - WITHDRAW ---
    deposits = 0.0
    withdraws = 0.0
    for t in out_trans:
        td = _get_trans_date(t)
        if not td or td.year != year:
            continue
        tt = _get_trans_type(t)
        amt = _get_trans_total(t)

        if tt == "DEPOSIT":
            deposits += amt
        elif tt == "WITHDRAW":
            withdraws += amt

    modal = deposits - withdraws
    return_ytd = (total_net_pl / modal) if modal else 0.0

    # Bulanan
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    monthly = []
    for m in range(1, 13):
        m_pl = 0.0
        m_fee = 0.0
        for d, v in daily_pl.items():
            if d.month == m:
                m_pl += v
                m_fee += daily_fee.get(d, 0.0)
        monthly.append((month_names[m - 1], m_pl, m_fee))

    # Mingguan: Week# mengikuti Excel WEEKNUM(date,2)
    def weeknum(d: date) -> int:
        return int(d.strftime("%W")) + 1

    max_week = weeknum(date(year, 12, 31))

    # --- Hari Trade: unique dates BELI/JUAL per week ---
    trade_days_by_week: Dict[int, set] = {w: set() for w in range(1, max_week + 1)}
    for t in out_trans:
        td = _get_trans_date(t)
        if not td or td.year != year:
            continue
        tt = _get_trans_type(t)
        if tt not in {"BELI", "JUAL"}:
            continue
        w = weeknum(td)
        trade_days_by_week.setdefault(w, set()).add(td)

    weekly_pl: Dict[int, float] = {w: 0.0 for w in range(1, max_week + 1)}
    weekly_fee: Dict[int, float] = {w: 0.0 for w in range(1, max_week + 1)}
    for d, v in daily_pl.items():
        w = weeknum(d)
        weekly_pl[w] = weekly_pl.get(w, 0.0) + v
        weekly_fee[w] = weekly_fee.get(w, 0.0) + daily_fee.get(d, 0.0)

    weekly = []
    for w in range(1, max_week + 1):
        weekly.append((w, weekly_pl.get(w, 0.0), weekly_fee.get(w, 0.0), len(trade_days_by_week.get(w, set()))))

    return {
        "year": year,
        "kpi": {
            "total_net_pl": total_net_pl,
            "total_fee": total_fee,
            "return_ytd": return_ytd,  # ratio, not percent
            "profit": profit,
            "loss": loss,
            "win_rate": win_rate,      # ratio
            "best_day": best_day,
            "worst_day": worst_day,
        },
        "monthly": monthly,
        "weekly": weekly,
    }

def save_table_to_txt(table_text: str, base_ts: str) -> str:
    os.makedirs("balance", exist_ok=True)
    path = os.path.join("balance", f"balance_{base_ts}.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(table_text)
    return path

def build_performance_overview_metrics(
        out: List[Dict[str, Any]], income_tax: Decimal, pl_amount: Decimal) -> List[Tuple[str, str]]:
    if not out:
        return [("Periode", "-")]

    dates = [_s(r["Tanggal"]) for r in out if r.get("Tanggal")]
    min_d = min(dates)
    max_d = max(dates)
    periode = f"{fmt_date_dd_mmm_yy(min_d)} s/d {fmt_date_dd_mmm_yy(max_d)}"

    hari_aktif = len(set(dates))
    total_trx = len(out)

    def count_type(t: str) -> int:
        return sum(1 for r in out if _upper(r["Transaksi"]) == t)

    def sum_total(t: str) -> Decimal:
        s = Decimal("0")
        for r in out:
            if _upper(r["Transaksi"]) == t:
                s += _to_decimal(r["Total"])
        return s

    deposit_cnt = count_type("DEPOSIT")
    withdraw_cnt = count_type("WITHDRAW")
    materai_cnt = count_type("MATERAI")
    beli_cnt = count_type("BELI")
    jual_cnt = count_type("JUAL")

    deposit_rp = sum_total("DEPOSIT")
    withdraw_rp = sum_total("WITHDRAW")
    materai_rp = sum_total("MATERAI")
    beli_rp = sum_total("BELI")
    jual_rp = sum_total("JUAL")

    modal_rp = deposit_rp - withdraw_rp

    total_fee = Decimal("0")
    for r in out:
        if _upper(r["Transaksi"]) in {"BELI", "JUAL"}:
            total_fee += _to_decimal(r.get("Total Fee"))

    # Operasional = Total Fee + Materai
    operasional = total_fee + materai_rp

    # Turnover gross = sum(price * lot * 100) BELI+JUAL
    turnover = Decimal("0")
    for r in out:
        if _upper(r["Transaksi"]) in {"BELI", "JUAL"}:
            turnover += _to_decimal(r.get("Harga")) * _to_decimal(r.get("Lot")) * Decimal("100")

    fee_ratio = Decimal("0")
    if turnover != 0:
        fee_ratio = (total_fee / turnover) * Decimal("100")

    fee_ratio_str = f"{fee_ratio.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)}%".replace(".", ",")

    # Net CF
    net_cf = (deposit_rp + jual_rp) - (withdraw_rp + beli_rp + materai_rp)

    return [
        ("Periode", periode),
        ("Hari aktif", fmt_int_id(hari_aktif)),
        ("Total transaksi", fmt_int_id(total_trx)),
        ("DEPOSIT (trx)", fmt_int_id(deposit_cnt)),
        ("DEPOSIT (Rp)", fmt_int_id(deposit_rp)),
        ("WITHDRAW (trx)", fmt_int_id(withdraw_cnt)),
        ("WITHDRAW (Rp)", fmt_int_id(withdraw_rp)),
        ("MODAL (Rp)", fmt_int_id(modal_rp)),
        ("Total Fee", fmt_int_id(total_fee)),
        ("Fee ratio", fee_ratio_str),
        ("Income Tax (Rp)", fmt_int_id(income_tax)),
        ("MATERAI (trx)", fmt_int_id(materai_cnt)),
        ("MATERAI (Rp)", fmt_int_id(materai_rp)),
        ("Operasional", fmt_int_id(operasional)),
        ("BELI (trx)", fmt_int_id(beli_cnt)),
        ("BELI (Rp)", fmt_int_id(beli_rp)),
        ("JUAL (trx)", fmt_int_id(jual_cnt)),
        ("JUAL (Rp)", fmt_int_id(jual_rp)),
        ("P/L (Rp)", fmt_int_id(pl_amount)),
        ("Turnover (gross)", fmt_int_id(turnover)),
        ("Net cashflow (Net CF)", fmt_int_id(net_cf)),
    ]

def add_overview_sheet(wb: Workbook, overview_metrics: List[Tuple[str, str]]) -> None:
    ws = wb.active
    ws.title = "Overview"

    headers = ["Metric", "Value"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    for m, v in overview_metrics:
        row = [m, v]
        ws.append(row)
        _update_widths(widths, row)

    ws.freeze_panes = "A2"

    # right align Value col
    right = Alignment(horizontal="right", vertical="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = right

    _apply_col_widths(ws, widths, pad=3, max_w=80)

def add_trade_performance_sheet(wb: Workbook, trade_metrics: List[Tuple[str, str]]) -> None:
    ws = wb.create_sheet("Trade Performance")

    headers = ["Trade KPI", "Value"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    for kpi, val in trade_metrics:
        row = [kpi, val]
        ws.append(row)
        _update_widths(widths, row)

    ws.freeze_panes = "A2"

    right = Alignment(horizontal="right", vertical="center")
    # Format numeric rows (Total realized P/L, Total profit, Total loss)
    for r in range(2, ws.max_row + 1):
        kpi = ws.cell(row=r, column=1).value
        vcell = ws.cell(row=r, column=2)
        if kpi in ("Total realized P/L", "Total profit", "Total loss"):
            vcell.value = _safe_float(vcell.value)
            vcell.number_format = FMT_MONEY_2
        vcell.alignment = right

    _apply_col_widths(ws, widths, pad=3, max_w=80)

def add_open_positions_sheet(wb: Workbook, open_positions: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Open Positions")

    headers = ["Ticker", "Open Lot", "Avg Cost", "Est. Cost (Rp)", "First Buy", "Last Buy"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    # body
    for r in open_positions:
        row = [
            r.get("Ticker", ""),
            _safe_int(r.get("Open Lot")),
            _safe_float(r.get("Avg Cost")),
            _safe_float(r.get("Est Cost")),
            r.get("First Buy"),
            r.get("Last Buy"),
        ]
        ws.append(row)
        _update_widths(widths, row)

    # total est cost
    total_est_cost = 0
    for r in open_positions:
        total_est_cost += _safe_int(r.get("Est Cost"))

    total_row = ["TOTAL", "", "", total_est_cost, "", ""]
    ws.append(total_row)
    _update_widths(widths, total_row)

    ws.freeze_panes = "A2"

    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # formats
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).alignment = left

        # Open Lot
        c = ws.cell(row=row_idx, column=2)
        c.number_format = FMT_INT
        c.alignment = right

        # Avg Cost, Est Cost
        for col in (3, 4):
            cc = ws.cell(row=row_idx, column=col)
            cc.value = _safe_float(cc.value)
            cc.number_format = FMT_MONEY_2
            cc.alignment = right

        # Dates
        for col in (5, 6):
            cd = ws.cell(row=row_idx, column=col)
            if cd.value:
                cd.number_format = FMT_DATE
            cd.alignment = left

    # bold TOTAL row
    tr = ws.max_row
    ws.cell(row=tr, column=1).font = Font(bold=True)
    ws.cell(row=tr, column=4).font = Font(bold=True)

    _apply_col_widths(ws, widths, pad=3, max_w=40)

def add_best_trades_sheet(wb: Workbook, best_trades: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Best Trades")

    headers = ["Ticker", "Realized P/L", "Avg Holding Period (d)"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    for r in best_trades:
        row = [
            r.get("Ticker", ""),
            _safe_float(r.get("Realized P/L", 0)),
            _safe_float(r.get("Avg Holding Period (d)", 0)),
        ]
        ws.append(row)
        _update_widths(widths, row)

    ws.freeze_panes = "A2"
    right = Alignment(horizontal="right", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = FMT_MONEY_2
        ws.cell(row=row_idx, column=2).alignment = right

        ws.cell(row=row_idx, column=3).number_format = "0.0"
        ws.cell(row=row_idx, column=3).alignment = right

    _apply_col_widths(ws, widths, pad=3, max_w=40)

def add_worst_trades_sheet(wb: Workbook, worst_trades: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Worst Trades")

    headers = ["Ticker", "Realized P/L", "Avg Holding Period (d)"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    for r in worst_trades:
        row = [
            r.get("Ticker", ""),
            _safe_float(r.get("Realized P/L", 0)),
            _safe_float(r.get("Avg Holding Period (d)", 0)),
        ]
        ws.append(row)
        _update_widths(widths, row)

    ws.freeze_panes = "A2"
    right = Alignment(horizontal="right", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = FMT_MONEY_2
        ws.cell(row=row_idx, column=2).alignment = right

        ws.cell(row=row_idx, column=3).number_format = "0.0"
        ws.cell(row=row_idx, column=3).alignment = right

    _apply_col_widths(ws, widths, pad=3, max_w=40)

def add_active_trading_sheet(
    wb: Workbook,
    active_daily: List[Dict[str, Any]],
    active_monthly: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Active Trading")

    # Section 1: Daily
    ws.append(["DAILY (Top Active Days by Turnover)"])
    ws["A1"].font = Font(bold=True)

    headers = ["Tanggal", "#Trx", "Turnover", "Net CF", "Realized P/L", "Fee"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 2, headers, header_fill, header_font)

    widths = _init_widths(headers)

    start_row = 3
    for r in active_daily:
        row = [
            r.get("Tanggal"),          # assume date or already formatted
            _safe_int(r.get("#Trx")),
            _safe_float(r.get("Turnover")),
            _safe_float(r.get("Net CF")),
            _safe_float(r.get("Realized P/L")),
            _safe_float(r.get("Fee")),
        ]
        ws.append(row)
        _update_widths(widths, row)

    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for r in range(start_row, ws.max_row + 1):
        # date
        cdate = ws.cell(row=r, column=1)
        if cdate.value:
            try:
                cdate.number_format = FMT_DATE
            except Exception:
                pass
        cdate.alignment = left

        # #trx
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=2).alignment = right

        # money columns
        for col in (3, 4, 5, 6):
            cell = ws.cell(row=r, column=col)
            cell.number_format = FMT_MONEY_2
            cell.alignment = right

    # blank + Section 2: Monthly
    blank_row = ws.max_row + 2
    ws.cell(row=blank_row, column=1, value="MONTHLY").font = Font(bold=True)

    header_row2 = blank_row + 1
    headers_m = ["Bulan", "#Trx", "Turnover", "Net CF", "Realized P/L", "Fee"]
    ws.append(headers_m)
    _style_header_row(ws, header_row2, headers_m, header_fill, header_font)

    # monthly rows
    for r in active_monthly:
        row = [
            r.get("Bulan"),
            _safe_int(r.get("#Trx")),
            _safe_float(r.get("Turnover")),
            _safe_float(r.get("Net CF")),
            _safe_float(r.get("Realized P/L")),
            _safe_float(r.get("Fee")),
        ]
        ws.append(row)
        _update_widths(widths, row)

    # format monthly rows
    for r in range(header_row2 + 1, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = left
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=2).alignment = right
        for col in (3, 4, 5, 6):
            cell = ws.cell(row=r, column=col)
            cell.number_format = FMT_MONEY_2
            cell.alignment = right

    ws.freeze_panes = "A3"
    _apply_col_widths(ws, widths, pad=3, max_w=80)

def add_gain_sheet(wb: Any, gain_blocks: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Gain")

    headers = [
        "Tanggal", "Transaksi", "Ticker", "Harga", "Lot",
        "Broker Fee", "VAT Broker Fee",
        "Exchange Fee", "VAT Exchange Fee",
        "Income Tax", "Total Fee", "Jumlah",
    ]
    ws.append(headers)

    widths = _init_widths(headers)
    header_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    summary_gain_font = Font(bold=True, color="FF0000")
    data_ranges: List[Tuple[int, int]] = []
    summary_rows: List[int] = []

    row_idx = 2

    for b in gain_blocks:
        rows: List[List[Any]] = b["rows"]
        totals: Dict[str, Decimal] = b["totals"]

        start_data = row_idx
        for r in rows:
            ws.append(r)
            _update_widths(widths, r)
            row_idx += 1
        end_data = row_idx - 1

        if end_data >= start_data:
            data_ranges.append((start_data, end_data))

        summary_row = [
            "", "", "", "", "",
            float(totals["broker_sum"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(totals["vat_broker_sum"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(totals["exch_sum"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(totals["vat_exch_sum"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            float(totals["tax_sum"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
            int(totals["fee_sum"].quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
            int(totals["gain"].quantize(Decimal("1"), rounding=ROUND_HALF_UP)),
        ]

        ws.append(summary_row)
        _update_widths(widths, summary_row)
        summary_rows.append(row_idx)
        row_idx += 1

        blank = [""] * len(headers)
        ws.append(blank)
        _update_widths(widths, blank)
        row_idx += 1

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    INT_COLS = {4, 5, 11, 12}
    MONEY_COLS = {6, 7, 8, 9, 10}

    for start_row, end_row in data_ranges:
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=len(headers)):
            c1 = row[0]
            c2 = row[1]
            c3 = row[2]

            if c1.value:
                c1.number_format = FMT_DATE
            c1.alignment = left

            c2.alignment = left
            c3.alignment = left

            for col_idx in INT_COLS:
                c = row[col_idx - 1]
                c.number_format = FMT_INT
                c.alignment = right

            for col_idx in MONEY_COLS:
                c = row[col_idx - 1]
                c.number_format = FMT_MONEY_2
                c.alignment = right

    for r_idx in summary_rows:
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=r_idx, column=col)
            c.alignment = right if col >= 4 else left

            if col in MONEY_COLS:
                c.number_format = FMT_MONEY_2
            elif col in (4, 5, 11):
                c.number_format = FMT_INT
            elif col == 12:
                c.number_format = FMT_GAIN
                c.font = summary_gain_font

    ws.freeze_panes = "A2"
    _apply_col_widths(ws, widths, pad=4, max_w=28)

def add_ticker_sheet(wb: Workbook, ticker_rows: List[Tuple[str, float, int]]) -> None:
    """
    Sheet 8: Ticker — (Ticker, Realized P/L, Qty) sorted desc.
    """
    ws = wb.create_sheet("Ticker")

    headers = ["Ticker", "Realized P/L", "Qty"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    _style_header_row(ws, 1, headers, header_fill, header_font)

    widths = _init_widths(headers)

    for ticker, pl, qty in ticker_rows:
        row = [ticker, _safe_float(pl), _safe_int(qty)]
        ws.append(row)
        _update_widths(widths, row)

    ws.freeze_panes = "A2"

    right = Alignment(horizontal="right", vertical="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = FMT_INT
        ws.cell(row=r, column=2).alignment = right
        ws.cell(row=r, column=3).number_format = "0"
        ws.cell(row=r, column=3).alignment = right

    _apply_col_widths(ws, widths, pad=3, max_w=30)

def add_performance_summary_sheet(wb: Workbook, summary: Dict[str, Any]) -> None:
    """
    Sheet 9/10: Performa 2025/2026 (layout KPI + Bulanan + Mingguan)
    """
    year = int(summary.get("year"))
    kpi = summary.get("kpi") or {}
    monthly = summary.get("monthly") or []
    weekly = summary.get("weekly") or []

    ws = wb.create_sheet(f"Performa {year}")

    # widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 12

    dark_blue = PatternFill("solid", fgColor="1F4E79")
    light_blue = PatternFill("solid", fgColor="D9E1F2")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    border = _border_thin()

    ws["A1"] = f"Ringkasan Performa {year}"
    ws["A1"].font = title_font

    # KPI header
    ws["A3"] = "KPI"
    ws["B3"] = "Nilai"
    for cell in ("A3", "B3"):
        ws[cell].fill = dark_blue
        ws[cell].font = white_font
        ws[cell].alignment = center
        ws[cell].border = border

    kpi_rows = [
        ("Total Net P/L (YTD)", kpi.get("total_net_pl", 0.0), "money"),
        ("Total Fee (YTD)", kpi.get("total_fee", 0.0), "money"),
        ("Return YTD (%)", kpi.get("return_ytd", 0.0), "pct"),
        ("Profit YTD", kpi.get("profit", 0.0), "money"),
        ("Loss YTD", kpi.get("loss", 0.0), "money"),
        ("Win Rate (%)", kpi.get("win_rate", 0.0), "pct"),
        ("Best Day (Rp)", kpi.get("best_day", 0.0), "money"),
        ("Worst Day (Rp)", kpi.get("worst_day", 0.0), "money"),
    ]
    start_row = 4
    for i, (label, val, kind) in enumerate(kpi_rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=label).alignment = left
        vcell = ws.cell(row=r, column=2, value=float(_safe_float(val)))
        vcell.alignment = right
        vcell.number_format = "0.00%" if kind == "pct" else FMT_INT
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = border

    # Bulanan
    ws["A13"] = "Bulanan"
    ws["A13"].font = bold

    ws["A14"] = "Bulan"
    ws["B14"] = "Net P/L"
    ws["C14"] = "Fee"
    for cell in ("A14", "B14", "C14"):
        ws[cell].fill = light_blue
        ws[cell].font = bold
        ws[cell].alignment = center
        ws[cell].border = border

    for i, (mn, mpl, mfee) in enumerate(monthly):
        r = 15 + i
        ws.cell(row=r, column=1, value=mn).alignment = left

        cpl = ws.cell(row=r, column=2, value=float(_safe_float(mpl)))
        cpl.number_format = FMT_INT
        cpl.alignment = right

        cfee = ws.cell(row=r, column=3, value=float(_safe_float(mfee)))
        cfee.number_format = FMT_INT
        cfee.alignment = right

        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border

    # Mingguan
    ws["H2"] = "Mingguan (Week#)"
    ws["H2"].font = bold
    ws.merge_cells("H2:K2")
    ws["H2"].alignment = center

    weekly_headers = ["Week#", "Net P/L", "Fee", "Hari Trade"]
    for j, h in enumerate(weekly_headers, start=8):
        cell = ws.cell(row=3, column=j, value=h)
        cell.fill = light_blue
        cell.font = bold
        cell.alignment = center
        cell.border = border

    for i, (w, wpl, wfee, wdays) in enumerate(weekly):
        r = 4 + i

        c0 = ws.cell(row=r, column=8, value=_safe_int(w))
        c0.alignment = center
        c0.border = border

        c1 = ws.cell(row=r, column=9, value=float(_safe_float(wpl)))
        c1.number_format = FMT_INT
        c1.alignment = right
        c1.border = border

        c2 = ws.cell(row=r, column=10, value=float(_safe_float(wfee)))
        c2.number_format = FMT_INT
        c2.alignment = right
        c2.border = border

        c3 = ws.cell(row=r, column=11, value=_safe_int(wdays))
        c3.number_format = "0"
        c3.alignment = center
        c3.border = border

def save_performance_excel(
    overview_metrics: List[Tuple[str, str]],
    trade_metrics: List[Tuple[str, str]],
    open_positions: List[Dict[str, Any]],
    best_trades: List[Dict[str, Any]],
    worst_trades: List[Dict[str, Any]],
    active_daily: List[Dict[str, Any]],
    active_monthly: List[Dict[str, Any]],
    gain_blocks: List[Dict[str, Any]],
    ticker_pl: List[Tuple[str, float, int]],
    base_ts: str,
    perf_summary_2025: Dict[str, Any],
    perf_summary_2026: Dict[str, Any],
) -> str:
    os.makedirs("balance", exist_ok=True)
    path = os.path.join("balance", f"performance_{base_ts}.xlsx")

    wb = Workbook()

    # Sheet 1-6 (konsisten: semua via function)
    add_overview_sheet(wb, overview_metrics)
    add_trade_performance_sheet(wb, trade_metrics)
    add_open_positions_sheet(wb, open_positions)
    add_best_trades_sheet(wb, best_trades)
    add_worst_trades_sheet(wb, worst_trades)
    add_active_trading_sheet(wb, active_daily, active_monthly)

    # -------------------------
    # Sheet 7: Gain (Journey BELI/JUAL)
    # -------------------------
    if gain_blocks:
        add_gain_sheet(wb, gain_blocks)

    # Sheet 8 — Ticker (Realized P/L by ticker)
    if ticker_pl:
        add_ticker_sheet(wb, ticker_pl)

    # Sheet 9 — Ringkasan Performa 2025
    if perf_summary_2025:
        add_performance_summary_sheet(wb, perf_summary_2025)

    # Sheet 10 — Ringkasan Performa 2026
    if perf_summary_2026:
        add_performance_summary_sheet(wb, perf_summary_2026)
    
    wb.save(path)
    return path

def generate_performance_excel(client: Client, base_ts: str) -> str:
    global _PNL_ACTIVE_CACHE
    _PNL_ACTIVE_CACHE = None

    out = fetch_mutasi_transaksi_for_display(client)

    income_tax = fetch_income_tax_from_pnl_summary(client)
    pl_amount = fetch_pl_from_pnl_summary(client)
    overview_metrics = build_performance_overview_metrics(out, income_tax, pl_amount)
    trade_metrics = build_trade_performance_metrics(client)
    open_positions = build_open_positions_from_trans(out)
    best_trades = build_best_trades_by_ticker(client, top_n=15)
    worst_trades = build_worst_trades_by_ticker(client, top_n=15)
    active_daily, active_monthly = build_active_trading_daily_monthly(out, client, top_n_days=15)
    gain_blocks = build_gain_blocks(client)
    ticker_pl = fetch_ticker_pl_from_pnl_summary(client)
    perf_2025 = build_yearly_performance_summary(client, out, 2025)
    perf_2026 = build_yearly_performance_summary(client, out, 2026)

    return save_performance_excel(
        overview_metrics,
        trade_metrics,
        open_positions,
        best_trades,
        worst_trades,
        active_daily,
        active_monthly,
        gain_blocks,
        ticker_pl,
        base_ts,
        perf_summary_2025=perf_2025,
        perf_summary_2026=perf_2026,
    )

def main() -> None:
    mode = (sys.argv[1].strip().lower() if len(sys.argv) > 1 else "")
    base_ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    client = create_supabase_client()

    if mode == "performance":
        # try:
        #     from openpyxl import Workbook
        #     from openpyxl.styles import Font, Alignment
        #     from openpyxl.utils import get_column_letter
        # except Exception as e:
        #     raise RuntimeError("openpyxl belum terinstall. Jalankan: pip install openpyxl") from e
        
        xlsx = generate_performance_excel(client, base_ts)
        print(f"Saved Performance Excel: {xlsx}")
        return

    out = fetch_mutasi_transaksi_for_display(client)

    if mode == "full":
        headers, rows = build_rows_full(out)
        right_cols = {
            "Harga", "Lot",
            "Broker Fee", "VAT Broker Fee", "Exchange Fee", "VAT Exchange Fee",
            "Income Fee", "Total Fee",
            "Total", "Saldo",
        }
        table = _render_table(headers, rows, right_cols=right_cols)
        txt = save_table_to_txt(table, base_ts)
        print(f"Saved TXT (FULL): {txt}")
        return

    headers, rows = build_rows_minimal(out)
    table = _render_table(headers, rows, right_cols={"Harga", "Lot", "Total", "Saldo"})
    txt = save_table_to_txt(table, base_ts)
    print(f"Saved TXT (MINIMAL): {txt}")


if __name__ == "__main__":
    main()
